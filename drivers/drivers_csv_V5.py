
'''
You Need to install 4-5 programs before you are able to run this script.
1. Python 3.4 or higher
2. PIP - if using python 3.4 pip should already be installed you can verify by typing "pip -V" into cmd prompt.
3. Pandas - goto cmd prompt type in: "pip install pandas" without quotes.
4. SQlite3 - if using python 3.4 or higher should already be included - necessary to create and operate on tabular data, and perform SQL like queries on data.
5. openpyxl - goto cmd prompt type in: "pip install openpyxl" without quotes.
6. Sublime Text - optional but makes editing and running scripts easier and may alleviate errors with file permissions.
'''

import sqlite3
import os
from datetime import datetime
from datetime import timedelta  
import pandas as pd
from pandas import DataFrame
import tkinter as tk
from tkinter import simpledialog
import re
from openpyxl import load_workbook

# Global Variables
USER_INP = ""
SHORT_DATE = ""
PATH = ""
TEMPLATE_PATH = ""


conn = None;
try:
    conn = sqlite3.connect(':memory:') # This allows the database to run in RAM, with no requirement to create a file.
    #conn = sqlite3.connect('dash_delivers.db')  # You can create a new database by changing the name within the quotes.
except Error as e:
    print(e)

c = conn.cursor() # The database will be saved in the location where your 'py' file is saved IF you did not choose the :memory: option
c.execute('''
  DROP TABLE IF EXISTS DRIVERS
    ''')
# Create table - DRIVERS from dash_output.csv - this must match the values and headers of the incoming CSV file.
c.execute('''CREATE TABLE IF NOT EXISTS DRIVERS
             ([Task_ID] INTEGER PRIMARY KEY, 
              [Complete_Before] DATE,
              [Completion_Time] DATE,
              [Order_ID] text, 
              [_Del Fee] DECIMAL(13,2), 
              [Total_Price] DECIMAL(13,2), 
              [Payment] text,
              [Tips] DECIMAL(13,2), 
              [Restaurant_Name] text, 
              [Agent_Name] text, 
              [Team_Name] text, 
              [Customer_Name] text, 
              [Customer_Address] text, 
              [Task_Status] text, 
              [Rating] text, 
              [Review] text)''')

conn.commit()

# read the csv file into our newly created SQL table DRIVERS
read_drivers = pd.read_csv (r'driver_outputs.csv', engine='python') # Use the python engine here to eliminate issues with UT-8
read_drivers.to_sql('DRIVERS', conn, if_exists='replace', index = False) # Insert the values from the csv file into the table 'DRIVERS'


# When reading the csv:
# - Place 'r' before the path string to read any special characters, such as '\'
# - Don't forget to put the file name at the end of the path + '.csv'
# - Before running the code, make sure that the column names in the CSV files match with the column names in the tables created and in the query below
# - If needed make sure that all the columns are in a TEXT format

#######################################################################################
# get_days_from_str(date_string):                                                     #
# Function to return the int value of a day from a time string.                       #
# Takes a string as an argument in the format YYYY-MM-DD HH:MM or %Y-%m-%d %H:%M      #
# and returns just the day value so if the string is 2020-05-28 10:23 function returns#
# 28 to the calling function. Returns current day if there is an error.               #
#######################################################################################
def get_days_from_str(date_string):
  try:
    #return datetime.strptime(date_string, "%Y-%m-%d").day
    return datetime.strptime(date_string, "%Y-%m-%d %H:%M").day
  except Exception as e:
    print(e, "Using %Y-%m-%d")
    return datetime.strptime(date_string, "%Y-%m-%d").day

  else:
    print("the date entered does not work")
    print(date_string)



#######################################################################################
# export_to_sheets():                                                                 #
# Function to export data fetched directly to workbook template                       #
#######################################################################################
def export_to_sheets(driver_name, data_frame, too_many_days, tip_array = []):

  # set file path
  filepath = TEMPLATE_PATH 
	# load demo.xlsx 
  wb=load_workbook(filepath)
	# get Sheet
  source=wb['Sheet1']

  # Enter data in tip array directly into known cells of the excel template.
  source['B2'] = tip_array[0]
  source['C2'] = tip_array[1]
  source['D2'] = tip_array[2]
  source['E2'] = tip_array[3]
  source['F2'] = tip_array[4]
  source['G2'] = tip_array[5]
  source['H2'] = tip_array[6]
  if too_many_days: # if required create another tip column.
    extra_day = SHORT_DATE + pd.DateOffset(days=7)   # SHORT_DATE + timedelta(days=7)
    source['I1'] = extra_day.strftime("%b.%d")
    source['I2'] = tip_array[7]

  row_jump = 15  # Sets the position of the data entry in the excel sheet, starting at row 15 for our template.
  col_jump = 1   # Since arrays are indexed starting at 0 we add 1 to match excel sheet values.
  if not data_frame.empty:
    current_day = get_days_from_str(data_frame.iat[0, 0])
  # Output the values of the drivers orders to excel sheet
    for i in range(len(data_frame)):  # from 0 to number of entries in our dataframe.
      if current_day != get_days_from_str(data_frame.iat[i, 0]):
        current_day = get_days_from_str(data_frame.iat[i, 0])
        row_jump += 1
      for j in range(0,8):
        cellref=source.cell(row=i+row_jump, column=j+col_jump)
        cellref.value=data_frame.iloc[i,j]
  try:
    #wb.save(PATH + '\\' + str(driver_name) + SHORT_DATE.strftime(" %b %d, %Y") + '.xlsx')
    wb.save(PATH + '\\' + str(driver_name) + datetime.now().strftime(" %b %d, %Y") + '.xlsx')
  except:
    print("unable to save output sheet of driver: " + str(driver_name))
    return
  
	# done
  return

#######################################################################################
# set_template_date():                                                                #
# Function to change the dates of the "_Driver Template.xlsx" template excel sheet and#
# save as a new dated template to be reused by program to create additional           #
# required.                                                                           #
#######################################################################################
def set_template_date():
  global TEMPLATE_PATH
  get_date_range() # prompts user for folder name and file appendature.
  get_start_date() # prompts user for the start of week date.
  temp_date = SHORT_DATE # sets temp variable from global SHORT_DATE which is the first day of extracted week.
  # set file path
  filepath ="_Driver Template.xlsx"
  # load filepath designated template 
  wb=load_workbook(filepath)
  # get Sheet
  source=wb['Sheet1']

  source['B1'] = temp_date.strftime("%b") + "." + str(temp_date.day)
  temp_date += timedelta(days=1)
  source['C1'] = temp_date.strftime("%b") + "." + str(temp_date.day)
  temp_date += timedelta(days=1)
  source['D1'] = temp_date.strftime("%b") + "." + str(temp_date.day)
  temp_date += timedelta(days=1)
  source['E1'] = temp_date.strftime("%b") + "." + str(temp_date.day)
  temp_date += timedelta(days=1)
  source['F1'] = temp_date.strftime("%b") + "." + str(temp_date.day)
  temp_date += timedelta(days=1)
  source['G1'] = temp_date.strftime("%b") + "." + str(temp_date.day)
  temp_date += timedelta(days=1)
  source['H1'] = temp_date.strftime("%b") + "." + str(temp_date.day)
  # save workbook
  TEMPLATE_PATH = PATH + '\\' + "driver_template_" + SHORT_DATE.strftime("%Y-%m-%d") + ".xlsx"
  wb.save(TEMPLATE_PATH)
  return

#######################################################################################
# get_date_range():                                                                   #
# Function to prompt user for required date range to append it to output files as     #
# required. * strips off special characters.                                          #
#######################################################################################
def get_date_range():
  ROOT = tk.Tk()
  ROOT.withdraw()
  global USER_INP
  global PATH

  USER_INP = simpledialog.askstring(title="Date Range",
                                  	  prompt="Input the date range to append to the end of each file name: \"Aug 12 - Aug 21 2019\"\nThis will also create a new folder of the same name in your current working directory.")
  USER_INP = re.sub('[^A-Za-z0-9\_\\-]+', '', USER_INP)

  path = os.getcwd()
  print ("The current working directory is %s" % path)
  path = path + '\\' + USER_INP 
  try:
    os.mkdir(path)
  except OSError:
    print ("Creation of the directory %s failed" % path)
  else:
    print ("Successfully created the directory %s " % path)
  PATH = path
  return

#######################################################################################
# get_start_date():                                                                   #
# Function to prompt user for the starting day of the week to be extracted it then    #
# uses this date to set the dates of the week for the printout excel sheet.           #
#######################################################################################
def get_start_date():
  ROOT = tk.Tk()
  ROOT.withdraw()
  global SHORT_DATE
  date_format = simpledialog.askstring(title="Start Date",
                                      prompt="Enter the start date in YYYY-MM-DD format")
  try:
    SHORT_DATE = datetime.strptime(date_format, "%Y-%m-%d")
  except:
    get_start_date() # continuous loop 
  return

#######################################################################################
# driver_pay(name):                                                                   #
# Function to pull required data from created SQL database, and export it into        #
# individual excel sheets for each driver.                                            #
#######################################################################################
def driver_pay(name):

  more_than_seven = False
  pay_days = [0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0]


  c.execute('''
  SELECT [Complete_Before], [Completion_Time], [Order_ID], [_Del Fee], [Total_Price], [Payment], [Restaurant_Name], [Tips]
  FROM DRIVERS
  WHERE Agent_Name == (?) 
  ORDER by SUBSTR([Complete_Before], 0, 11)
    ''', (name,))

  df = DataFrame(c.fetchall(), columns=['Complete_Before', 'Completion_Time', 'Order_ID', '_Del Fee', 'Total_Price', 'Payment', 'Restaurant_Name','Tips'])

  c.execute('''
  SELECT SUBSTR([Complete_Before], 0, 11) AS stripped_time, SUM([Tips])
  FROM DRIVERS
  WHERE Agent_Name == (?) AND [Payment] != \'CANCELLED\'
  GROUP BY stripped_time
    ''', (name,))

  pf = c.fetchall()
  
  for pay_day in pf:
    try:
      a = datetime.strptime(pay_day[0], "%Y-%m-%d") - SHORT_DATE
      if a.days == 0:
        pay_days[0] = float(pay_day[1])
      elif a.days == 1:
        pay_days[1] = float(pay_day[1])
      elif a.days == 2:
        pay_days[2] = float(pay_day[1])
      elif a.days == 3:
        pay_days[3] = float(pay_day[1])
      elif a.days == 4:
        pay_days[4] = float(pay_day[1])
      elif a.days == 5:
        pay_days[5] = float(pay_day[1])
      elif a.days == 6:
        pay_days[6] = float(pay_day[1])
      elif a.days == 7:
        pay_days[7] = float(pay_day[1])
        print("The inputs total up to more than seven days, adding in an additional column in Tips for: " + name)
        more_than_seven = True
      else:
        print("The input dates do not match the first day of week entered. Exiting")
        exit()
    except Exception as e:
      print(name)
      raise

      

  export_to_sheets(name, df, more_than_seven, pay_days)

  return
   ###############################################################
   # End of Function to pull out data and export to sheets       #
   ###############################################################


# Prompts user for starting day of the week, and creates a template based on those days.
set_template_date()

# Select all Drivers in the .csv so we can run a for loop with each name.
c.execute('''
	SELECT [Agent_Name]
	FROM DRIVERS
	GROUP BY [Agent_Name]
		 ''')

driver_run = c.fetchall()

for driver in driver_run:
  driver_pay(driver[0]) # using the first column, even though the query returns only one column it needs to be specified.

#driver_pay("Wesley Mann") 
# Drop the table that was created.
c.execute('''
	DROP TABLE IF EXISTS DRIVERS
		''')

# Close the database connection
conn.close()



'''
You Need to install 4-5 programs before you are able to run this script.
1. Python 3.4 or higher
2. PIP - if using python 3.4 pip should already be installed you can verify by typing "pip -V" into cmd prompt.
3. Pandas - goto cmd prompt type in: "pip install pandas" without quotes.
4. SQlite3 - if using python 3.4 or higher should already be included - necessary to create and operate on tabular data, and perform SQL like queries on data.
5. openpyxl - goto cmd prompt type in: "pip install openpyxl" without quotes.
6. tkcalendar - goto cmd prompt type in: "pip install tkcalendar" without quotes.
7. Sublime Text - optional but makes editing and running scripts easier and may alleviate errors with file permissions.
'''

import sqlite3
#from dateutil.parser import parse
import pandas as pd
from pandas import DataFrame
import tkinter as tk
import re
from openpyxl import load_workbook
import os
from datetime import datetime, timedelta 
from tkinter import simpledialog, ttk
from tkcalendar import Calendar, DateEntry



# Global Variables
USER_INP = ""
SHORT_DATE = None
FOLDER_DATE = None
PATH = ""
TEMPLATE_PATH = ""


conn = None;
try:
    conn = sqlite3.connect(':memory:') # This allows the database to run in RAM, with no requirement to create a file.
    #conn = sqlite3.connect('dash_delivers.db')  # You can create a new database by changing the name within the quotes.
    #print(sqlite3.version)
except Error as e:
    print(e)



c = conn.cursor() # The database will be saved in the location where your 'py' file is saved IF you did not choose the :memory: option

# Create table - DRIVERS from dash_output.csv - this must match the values and headers of the incoming CSV file.
c.execute('''CREATE TABLE IF NOT EXISTS DRIVERS
             ([generated_id] INTEGER PRIMARY KEY, 
              [Source.Name] text, 
              [Order ID] text, 
              [Type] text, 
              [Outcome] text, 
              [First Name] text,
              [Last Name] text, 
              [Address] text, 
              [Subtotal] DECIMAL(13,2), 
              [Delivery fee] DECIMAL(13,2), 
              [Tip amount - gross] DECIMAL(13,2), 
              [Total taxes] DECIMAL(13,2), 
              [Total] DECIMAL(13,2), 
              [Payment Method] text, 
              [Fulfillment date (YYYY-MM-DD)] date, 
              [Fulfillment time] text, 
              [Confirmed date (YYYY-MM-DD)] text,
              [Confirmed time] text, 
              [Placed date (YYYY-MM-DD)] text, 
              [Placed time (hh:mm)] text, 
              [Service fees on subtotal] DECIMAL(13,2) )''')

# Create our own custom table - DASH_DATA to store the sums of the order values for the week.
c.execute('''CREATE TABLE IF NOT EXISTS DASH_DATA
             ([generated_id] INTEGER PRIMARY KEY, 
              [restaurant_name] text, 
              [sales_total] DECIMAL(13,2),  
              [delivery_debit_total] DECIMAL(13,2),
              [delivery_cash_total] DECIMAL(13,2), 
              [delivery_fee_debit] DECIMAL(13,2), 
              [delivery_fee_cash] DECIMAL(13,2), 
              [service_fee_total_cash] DECIMAL(13,2), 
              [service_fee_total_debit] DECIMAL(13,2), 
              [pickup_online] DECIMAL(13,2), 
              [pickup_instore] DECIMAL(13,2), 
              [pickup_online_tips] DECIMAL(13,2))''')             
conn.commit()

read_drivers = pd.read_csv (r'dash_output.csv', engine='python')
read_drivers.to_sql('DRIVERS', conn, if_exists='replace', index = False) # Insert the values from the csv file into the table 'DRIVERS'


# When reading the csv:
# - Place 'r' before the path string to read any special characters, such as '\'
# - Don't forget to put the file name at the end of the path + '.csv'
# - Before running the code, make sure that the column names in the CSV files match with the column names in the tables created and in the query below
# - If needed make sure that all the columns are in a TEXT format

#######################################################################################
# return_value(fetch_return):                                                         #
# Function to return the float value of a string returned by an SQL query, returns    #
# the float value of the string, or 0.0 if the value is NULL or "not a number" NAN.   #
#######################################################################################
def return_value(fetch_return):
  try:
    for item in fetch_return[0]:
      return float(item)
  except:
      return 0.0


#######################################################################################
# export_to_sheets():                                                                 #
# Function to export data fetched directly to workbook template                       #
#######################################################################################
def export_to_sheets(name, total, pickup_online, pickup_instore, pickup_online_tips, 
                     delivery_debit_total, delivery_cash_total, delivery_fee_debit, 
                     delivery_fee_cash, service_fee_total_cash, service_fee_total_debit):

  # set file path
  filepath = TEMPLATE_PATH 
  # load demo.xlsx 
  wb=load_workbook(filepath)
  # get Sheet
  source=wb['Sheet1']

  # Enter data in tip array directly into known cells of the excel template.
  source['D18'] = pickup_online
  source['E18'] = pickup_instore
  source['D21'] = pickup_online_tips
  source['F18'] = delivery_debit_total
  source['G18'] = delivery_cash_total
  source['B34'] = delivery_fee_debit
  source['B35'] = delivery_fee_cash
  source['G20'] = service_fee_total_cash
  source['F20'] = service_fee_total_debit

  try:
    wb.save(PATH + '\\' + str(name) + FOLDER_DATE.strftime(" %b %d, %Y") + '.xlsx')
    #wb.save(PATH + '\\' + str(driver_name) + datetime.now().strftime(" %b %d, %Y") + '.xlsx')
  except:
    print("unable to save output sheet of driver: " + str(name))
    return
  
  # done
  return

#######################################################################################
# get_date_range():                                                                   #
# Function to prompt user for required date range to append it to output files as     #
# required. * strips off special characters.                                          #
#######################################################################################
def get_date_range():
  root = tk.Tk()
  root.withdraw()
  global USER_INP
  global PATH
  date1 = datetime.now()
  USER_INP = simpledialog.askstring(title="Date Range",
                                    prompt="Input the date range of the working week: \""
                                     + (date1 - timedelta(days=7)).strftime("%b %d - ")
                                     + date1.strftime("%b %d %Y") 
                                     + "\"\nThis will create a new folder of the same name in your current working directory.")
  if USER_INP is not None:
    USER_INP = re.sub('[^A-Za-z0-9\_\\-]+', '_', USER_INP)
  else:
    print ("Creation of the directory cancelled. Exiting")
    conn.close()
    exit()

  path = os.getcwd()
  print ("The current working directory is %s" % path)
  path = path + '\\' + USER_INP 
  try:
    os.mkdir(path)
  except OSError:
    print ("Creation of the directory %s failed" % path)
    conn.close()
    exit()
  else:
    print ("Successfully created the directory %s " % path)
  PATH = path
  return


#######################################################################################
# calendar_gui():                                                                     #
# Function to prompt user for required dates to append it to output files as          #
# required.                                                                           #
#######################################################################################
def calendar_gui():
  def on_closing():
    close_windows()
  def close_windows():
    global FOLDER_DATE
    try:
      if SHORT_DATE is None:
        set_start_date() # continuous loop until proper date entered.
      elif FOLDER_DATE is None:
        print ("Entry of append date cancelled. Using todays date")
        FOLDER_DATE = datetime.now()
        root.destroy()
      else:
        root.destroy()
        return
    except Exception as e:
      raise e

  def set_start_date():
    def print_sel():
      global SHORT_DATE
      SHORT_DATE = cal.selection_get()
      #print((cal.selection_get())
      top.destroy()

    top = tk.Toplevel(root)

    cal = Calendar(top,
                   font="Arial 14", selectmode='day',
                   cursor="hand1", year=datetime.now().year, month=datetime.now().month, day=datetime.now().day)
    cal.pack(fill="both", expand=True)
    ttk.Button(top, text="ok", command=print_sel).pack()

  def set_folder_date():
    def print_sel():
      global FOLDER_DATE
      FOLDER_DATE = cal.selection_get()
      top.destroy()

    top = tk.Toplevel(root)

    cal = Calendar(top,
                   font="Arial 14", selectmode='day',
                   cursor="hand1", year=datetime.now().year, month=datetime.now().month, day=datetime.now().day)
    cal.pack(fill="both", expand=True)
    ttk.Button(top, text="ok", command=print_sel).pack()


  root = tk.Tk()
  root.protocol("WM_DELETE_WINDOW", on_closing)
  s = ttk.Style(root)
  s.theme_use('clam')
  root.title('Dash - enter work week dates')
  #root.geometry("350x180+300+300") #Width x Height

  w = 350 # width for the Tk root
  h = 180 # height for the Tk root

  # get screen width and height
  ws = root.winfo_screenwidth() # width of the screen
  hs = root.winfo_screenheight() # height of the screen

  # calculate x and y coordinates for the Tk root window
  x = (ws/2) - (w/2)
  y = (hs/2) - (h/2)

  # set the dimensions of the screen 
  # and where it is placed
  root.geometry('%dx%d+%d+%d' % (w, h, x, y))

  ttk.Button(root, text='Set Start Date', command=set_start_date).pack(padx=10, pady=10)
  ttk.Button(root, text='Set Folder Date', command=set_folder_date).pack(padx=10, pady=10)
  ttk.Button(root, text='Close', command=close_windows).pack(padx=10, pady=10)
    #calendar_view()
  root.mainloop()

  return

#######################################################################################
# set_template_date():                                                                #
# Function to change the dates of the "_Driver Template.xlsx" template excel sheet and#
# save as a new dated template to be reused by program to create additional           #
# required.                                                                           #
#######################################################################################
def set_template_date():
  global TEMPLATE_PATH
  calendar_gui()
  get_date_range() # prompts user for folder name and file appendature.

  temp_date = SHORT_DATE # sets temp variable from global SHORT_DATE which is the first day of extracted week.
  # set file path
  filepath ="template_settlement.xlsx"
  # load filepath designated template 
  wb=load_workbook(filepath)
  # get Sheet
  source=wb['Sheet1']

  source['B4'] = temp_date.strftime("%b %d - ") + (temp_date + timedelta(days=7)).strftime("%b %d %Y")

  # save workbook
  TEMPLATE_PATH = PATH + '\\' + "set_template_" + SHORT_DATE.strftime("%Y-%m-%d") + ".xlsx"
  wb.save(TEMPLATE_PATH)
  return

#######################################################################################
# total_sales(name):                                                                  #
# Function to pull required data from created SQL database, and store it in the newly #
# created table DASH_DATA.                                                            #
#######################################################################################
def total_sales(name):
   # Pull total With Taxes
  c.execute('''
	SELECT SUM(DRIVERS.[Total]) 
	FROM DRIVERS 
	WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?)
	GROUP BY DRIVERS.[Source.Name]''', (name,))

  total = return_value(c.fetchall())

  c.execute('''
  SELECT SUM(DRIVERS.[Tip amount - gross]) 
  FROM DRIVERS 
  WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?)
  GROUP BY DRIVERS.[Source.Name]''', (name,))

  total_tips = return_value(c.fetchall())

  total = total - total_tips  # Removes tips off of total for online purchases - repetitive code... work to remove floating points and duplicate code.
   # END Pull total
   ###########################################################################################################################
   # Pull pickup online total with taxes 
  c.execute('''
	SELECT SUM(DRIVERS.[Total]) 
	FROM DRIVERS 
	WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'pickup\' AND (DRIVERS.[Payment Method] == \'ONLINE\' OR DRIVERS.[Payment Method] == \'PayPal\')
	GROUP BY DRIVERS.[Source.Name]''', (name,))

  pickup_online = return_value(c.fetchall())

  c.execute('''
  SELECT SUM(DRIVERS.[Tip amount - gross]) 
  FROM DRIVERS 
  WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'pickup\' AND (DRIVERS.[Payment Method] == \'ONLINE\' OR DRIVERS.[Payment Method] == \'PayPal\')
  GROUP BY DRIVERS.[Source.Name]''', (name,))

  pickup_online_tips = return_value(c.fetchall())

  pickup_online = pickup_online - pickup_online_tips # Removes tips off of pickup_total for online purchases - repetitive code work to remove floating points and duplicate code.
  # END Pull Pickup online Total
  ###########################################################################################################################
  #Pull pickup instore total with taxes 

  c.execute('''
  SELECT SUM(DRIVERS.[Total]) 
  FROM DRIVERS 
  WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'pickup\' AND DRIVERS.[Payment Method] != \'ONLINE\' AND DRIVERS.[Payment Method] != \'PayPal\'
  GROUP BY DRIVERS.[Source.Name]''', (name,))

  pickup_instore = return_value(c.fetchall())

  c.execute('''
  SELECT SUM(DRIVERS.[Tip amount - gross]) 
  FROM DRIVERS 
  WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'pickup\' AND DRIVERS.[Payment Method] != \'ONLINE\' AND DRIVERS.[Payment Method] != \'PayPal\'
  GROUP BY DRIVERS.[Source.Name]''', (name,))

  pickup_instore_tips = return_value(c.fetchall())

  pickup_instore = pickup_instore - pickup_instore_tips # Removes tips off of pickup_total for online purchases - repetitive code work to remove floating points and duplicate code.

  # END Pull Pickup Instore Total
  ###########################################################################################################################
  # Pull delivery debit total with taxes
  c.execute('''
	SELECT SUM(DRIVERS.[Total]) 
	FROM DRIVERS 
	WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'delivery\' AND (DRIVERS.[Payment Method] == \'CARD\' OR DRIVERS.[Payment Method] == \'ONLINE\' OR DRIVERS.[Payment Method] == \'PayPal\')
	GROUP BY DRIVERS.[Source.Name]''', (name,))

  delivery_debit_total = return_value(c.fetchall())

  c.execute('''
  SELECT SUM(DRIVERS.[Tip amount - gross]) 
  FROM DRIVERS 
  WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'delivery\' AND (DRIVERS.[Payment Method] == \'CARD\' OR DRIVERS.[Payment Method] == \'ONLINE\')
  GROUP BY DRIVERS.[Source.Name]''', (name,))
 
  delivery_debit_tips = return_value(c.fetchall())

  delivery_debit_total = delivery_debit_total - delivery_debit_tips
    # END pull delivery debit total
    ###########################################################################################################################
    # Pull delivery cash total with taxes
  c.execute('''
	SELECT SUM(DRIVERS.[Total]) 
	FROM DRIVERS 
	WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'delivery\' AND DRIVERS.[Payment Method] == \'CASH\'
	GROUP BY DRIVERS.[Source.Name]''', (name,))

  delivery_cash_total = return_value(c.fetchall())

  c.execute('''
  SELECT SUM(DRIVERS.[Tip amount - gross]) 
  FROM DRIVERS 
  WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'delivery\' AND DRIVERS.[Payment Method] == \'CASH\'
  GROUP BY DRIVERS.[Source.Name]''', (name,))

  delivery_cash_tips = return_value(c.fetchall())

  delivery_cash_total = delivery_cash_total - delivery_cash_tips

    # END pull delivery cash total
    ###########################################################################################################################
    # Pull delivery fee total (debit)
  c.execute('''
	SELECT SUM(DRIVERS.[Delivery fee]) 
	FROM DRIVERS 
	WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'delivery\' AND (DRIVERS.[Payment Method] == \'CARD\' OR DRIVERS.[Payment Method] == \'ONLINE\' OR DRIVERS.[Payment Method] == \'PayPal\')
	GROUP BY DRIVERS.[Source.Name]''', (name,))

  delivery_fee_debit = return_value(c.fetchall())

    # END pull delivery fee total

   # Pull delivery fee total (cash)
  c.execute('''
	SELECT SUM(DRIVERS.[Delivery fee]) 
	FROM DRIVERS 
	WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'delivery\' AND DRIVERS.[Payment Method] == \'CASH\'
	GROUP BY DRIVERS.[Source.Name]''', (name,))

  delivery_fee_cash = return_value(c.fetchall())
    # END pull delivery fee total (cash)

    # Pull service fee total (cash)
  c.execute('''
	SELECT SUM(DRIVERS.[Service fees on subtotal]) 
	FROM DRIVERS 
	WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'delivery\' AND DRIVERS.[Payment Method] == \'CASH\'
	GROUP BY DRIVERS.[Source.Name]''', (name,))

  service_fee_total_cash = return_value(c.fetchall())
    # END pull service fee total


        # Pull service fee total (card)
  c.execute('''
	SELECT SUM(DRIVERS.[Service fees on subtotal]) 
	FROM DRIVERS 
	WHERE DRIVERS.[Outcome] == \'accepted\' AND DRIVERS .[Source.Name] == (?) AND DRIVERS.[Type] == \'delivery\' AND (DRIVERS.[Payment Method] == \'CARD\' OR DRIVERS.[Payment Method] == \'ONLINE\' OR DRIVERS.[Payment Method] == \'PayPal\')
	GROUP BY DRIVERS.[Source.Name]''', (name,))

  service_fee_total_debit = return_value(c.fetchall())
    # END pull service fee total

    # insert into the created table DASH_DATA the values pulled from the table DASH_DRIVERS
  c.execute('''INSERT INTO DASH_DATA (restaurant_name, sales_total, pickup_online, pickup_instore, pickup_online_tips,delivery_debit_total, delivery_cash_total, delivery_fee_debit, delivery_fee_cash, service_fee_total_cash, service_fee_total_debit) VALUES ((?),(?),(?),(?),(?),(?),(?),(?),(?),(?),(?))''',
    (name, total, pickup_online, pickup_instore, pickup_online_tips, delivery_debit_total, delivery_cash_total, delivery_fee_debit, delivery_fee_cash, service_fee_total_cash, service_fee_total_debit))
  
  export_to_sheets(name,
                   total,
                   pickup_online, 
                   pickup_instore, 
                   pickup_online_tips, 
                   delivery_debit_total, 
                   delivery_cash_total, 
                   delivery_fee_debit, 
                   delivery_fee_cash, 
                   service_fee_total_cash, 
                   service_fee_total_debit)
   
  return
   ###############################################################
   # End of Function to pull out data                            #
   ###############################################################
 
# Prompts user for starting day of the week, and creates a template based on those days.
set_template_date()
# Select all Restaurants in the .csv so we can run a for loop with each name.
c.execute('''
	SELECT DRIVERS.[Source.Name]
	FROM DRIVERS
	GROUP BY DRIVERS.[Source.Name]
		 ''')

restaurant_run = c.fetchall()

for restaurant in restaurant_run:
	total_sales(restaurant[0]) # using the first column, even though the query returns only one column it needs to be specified.


c.execute('''
	SELECT [restaurant_name], [sales_total], [pickup_online_tips], [pickup_online], [pickup_instore], [delivery_debit_total], [delivery_cash_total], [delivery_fee_debit], [delivery_fee_cash], [service_fee_total_debit], [service_fee_total_cash]
	FROM DASH_DATA 
		 ''')

#df = DataFrame(c.fetchall())
df = DataFrame(c.fetchall(), columns=['Source.Name', 'Subtotal', 'Pickup Tips', 'Pickup (Online)', 'Pickup (Instore)', 'Delivery Total (Debit)', 'Delivery Total (Cash)', 'Delivery Fee (Debit)', 'Delivery Fee (Cash)', 'Service Fee Total (debit)', 'Service Fee Total (cash)'])
print (df) 


#get_date_range()
#export_to_sheets()
#df.to_sql('DRIVERS', conn, if_exists='append', index = False) # Insert the values from the INSERT QUERY into the table 'DAILY_STATUS'

try:
	export_csv = df.to_csv (r'export_list.csv', index = None, header=True) # Export the results to a CSV. Make sure to adjust the path name
except PermissionError:
	print("export_list.csv is open, cannot save output.")
# Don't forget to add '.csv' at the end of the path (as well as r at the beg to address special characters)

c.execute('''
	DROP TABLE IF EXISTS DRIVERS
		''')
c.execute('''
	DROP TABLE IF EXISTS DASH_DATA
		''')


conn.close()

